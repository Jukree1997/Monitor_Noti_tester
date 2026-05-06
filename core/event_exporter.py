"""Event-log export — writes a snapshot of a session's recorded events into
a single ``report.xlsx`` workbook.

The runner emits structured events via ``event_recorded(dict, str)`` (event +
status). The UI (or a headless harness) accumulates those into a buffer; this
module turns that buffer into a multi-sheet report.

Sheets in ``report.xlsx`` (in order):
  * **Lines** — every line crossing event, raw.
  * **Zones** — every zone enter/exit/overstay event, with pairing + dwell.
  * **Summary** — whole-session aggregates per line / zone / area.
  * **Hourly_Summary** — per-(date,hour) blocks of the same aggregates,
    with dwell clipped to hour boundaries.
  * **Daily_Summary** — per-day rankings: peak hour for entrances, exits,
    longest visit, etc.
  * **Metadata** — export context (project, model, source, quality hints).

Public entry: :func:`export_events`. Pure functions — no Qt, no UI imports.
"""
from __future__ import annotations
import os
import re
import time
from typing import Iterable


# Event-type vocabulary used by the runner. Kept here so the exporter is the
# single source of truth for "is this a line / zone / area row?".
LINE_EVENT_TYPES = ("line_in", "line_out", "line_cross_in", "line_cross_out")
ZONE_PER_OBJECT_TYPES = ("zone_enter", "zone_exit", "zone_overstay")
AREA_EVENT_TYPES = ("stuck",)

EVENT_LABEL = {
    "line_in": "IN",
    "line_out": "OUT",
    "line_cross_in": "IN",
    "line_cross_out": "OUT",
    "zone_enter": "ENTER",
    "zone_exit": "EXIT",
    "zone_overstay": "OVERSTAY",
    "stuck": "STUCK",
}


def _safe_name(s: str) -> str:
    """Sanitise a project name into a directory-safe slug."""
    cleaned = re.sub(r"[^A-Za-z0-9_-]+", "_", s or "").strip("_")
    return cleaned or "project"


def _iso(ts: float) -> str:
    if not ts:
        return ""
    return time.strftime("%Y-%m-%dT%H:%M:%S", time.localtime(ts))


def _time_local(ts: float) -> str:
    if not ts:
        return ""
    return time.strftime("%H:%M:%S", time.localtime(ts))


def _format_dwell(seconds: float | None, decimals: int = 1) -> str:
    if seconds is None:
        return ""
    return f"{seconds:.{decimals}f}"


def compute_pairings(buffer: list[dict],
                     currently_tracked_ids: set[int],
                     export_time: float) -> dict[int, dict]:
    """Pair zone enters with the next zone exit for the same (region, object).

    Returns ``{buffer_index: {"pairing": str, "dwell_sec": float|None}}``.

    ``pairing`` values:
      * ``"matched"``      — enter has a paired exit later in the buffer
      * ``"still_inside"`` — enter has no exit; object is still being tracked
      * ``"lost_id"``      — enter has no exit; object_id no longer tracked
      * ``"n/a"``          — exit / overstay rows; ``dwell_sec`` is informational

    Multiple in/out cycles per (region, object) are handled by walking events
    in time order with a single open-enter cursor (no nested enters possible
    on the same key — the zone manager toggles per frame).
    """
    pairings: dict[int, dict] = {}

    by_key: dict[tuple, list[tuple[int, str, float]]] = {}
    for i, ev in enumerate(buffer):
        et = ev.get("event_type")
        if et in ZONE_PER_OBJECT_TYPES:
            key = (str(ev.get("region_id", "")), int(ev.get("object_id", 0)))
            by_key.setdefault(key, []).append(
                (i, et, float(ev.get("timestamp", 0.0))))

    for events in by_key.values():
        events.sort(key=lambda e: e[2])
        open_idx: int | None = None
        open_ts: float | None = None

        def _close_open_unpaired():
            """Mark the dangling open enter as still_inside or lost_id."""
            nonlocal open_idx, open_ts
            if open_idx is None:
                return
            obj_id = int(buffer[open_idx].get("object_id", 0))
            in_tracker = obj_id in currently_tracked_ids
            pairing = "still_inside" if in_tracker else "lost_id"
            dwell = (export_time - open_ts) if open_ts is not None else None
            pairings[open_idx] = {"pairing": pairing, "dwell_sec": dwell}
            open_idx = None
            open_ts = None

        for buf_idx, et, ts in events:
            if et == "zone_enter":
                if open_idx is not None:
                    # Two enters with no exit between — first one is orphan.
                    _close_open_unpaired()
                open_idx = buf_idx
                open_ts = ts
            elif et == "zone_exit":
                if open_idx is not None:
                    dwell = ts - (open_ts or ts)
                    pairings[open_idx] = {"pairing": "matched", "dwell_sec": dwell}
                    pairings[buf_idx] = {"pairing": "n/a", "dwell_sec": dwell}
                    open_idx = None
                    open_ts = None
                else:
                    pairings[buf_idx] = {"pairing": "n/a", "dwell_sec": None}
            elif et == "zone_overstay":
                # Overstay rides on the same open-enter cursor; it does NOT
                # close it (object is still inside, same enter still pending).
                if open_idx is not None and open_ts is not None:
                    dwell = ts - open_ts
                    pairings[buf_idx] = {"pairing": "n/a", "dwell_sec": dwell}
                else:
                    pairings[buf_idx] = {"pairing": "n/a", "dwell_sec": None}

        # End of stream for this key — if an enter is still open, it's orphan.
        _close_open_unpaired()

    return pairings


def _ev_status(ev: dict) -> str:
    return str(ev.get("status", ""))


def _ev_quality(ev: dict) -> str:
    return str(ev.get("quality", "ok"))


def _ev_mode(ev: dict) -> str:
    return str(ev.get("mode", ""))


LINES_HEADERS = ["timestamp_iso", "time_local", "event", "line_name",
                 "object_id", "class", "status", "mode"]
ZONES_HEADERS = ["timestamp_iso", "time_local", "event", "scope", "region",
                 "object_id", "class", "dwell_sec", "quality", "pairing",
                 "status", "mode", "details"]


def build_lines_rows(buffer: list[dict]) -> list[list]:
    """Pure row builder shared by the CSV and XLSX writers — keeps schema
    drift impossible since both formats emit the same data."""
    rows: list[list] = []
    for ev in buffer:
        et = ev.get("event_type", "")
        if et not in LINE_EVENT_TYPES:
            continue
        ts = float(ev.get("timestamp", 0.0))
        rows.append([
            _iso(ts), _time_local(ts),
            EVENT_LABEL.get(et, et),
            ev.get("region_name", ""),
            ev.get("object_id", 0),
            ev.get("class_name", ""),
            _ev_status(ev),
            _ev_mode(ev),
        ])
    return rows


def build_zones_rows(buffer: list[dict],
                     pairings: dict[int, dict]) -> list[list]:
    rows: list[list] = []
    for i, ev in enumerate(buffer):
        et = ev.get("event_type", "")
        if et in ZONE_PER_OBJECT_TYPES:
            scope = "object"
        elif et in AREA_EVENT_TYPES:
            scope = "area"
        else:
            continue
        pairing_info = pairings.get(i, {"pairing": "n/a", "dwell_sec": None})
        ts = float(ev.get("timestamp", 0.0))
        rows.append([
            _iso(ts), _time_local(ts),
            EVENT_LABEL.get(et, et),
            scope,
            ev.get("region_name", ""),
            ev.get("object_id", 0),
            ev.get("class_name", ""),
            _format_dwell(pairing_info.get("dwell_sec")),
            _ev_quality(ev),
            pairing_info.get("pairing", "n/a"),
            _ev_status(ev),
            _ev_mode(ev),
            ev.get("details", ""),
        ])
    return rows


def _summary_for_lines(buffer: list[dict]) -> list[dict]:
    """Aggregate per-line counts; one row per line_name encountered."""
    by_line: dict[str, dict] = {}
    for ev in buffer:
        et = ev.get("event_type", "")
        if et not in LINE_EVENT_TYPES:
            continue
        name = ev.get("region_name", "") or "(unnamed)"
        bucket = by_line.setdefault(name, {
            "in_count": 0, "out_count": 0, "objs": set()})
        if et in ("line_in", "line_cross_in"):
            bucket["in_count"] += 1
        else:
            bucket["out_count"] += 1
        bucket["objs"].add(int(ev.get("object_id", 0)))

    rows: list[dict] = []
    for name, b in sorted(by_line.items()):
        rows.append({
            "type": "line", "name": name,
            "in_count": b["in_count"], "out_count": b["out_count"],
            "unique_objects": len(b["objs"]),
        })
    return rows


def _summary_for_zones(buffer: list[dict],
                       pairings: dict[int, dict]) -> list[dict]:
    """Aggregate per-zone enter/exit counts and dwell statistics.

    Dwell stats include both completed (``pairing=matched``) and in-progress
    (``pairing=still_inside``) enters — the latter use ``now - enter_time``,
    so a long-parked object that hasn't left yet still shows up in
    ``max_dwell``. ``session_start`` enters (object already in the zone at
    detection start) are included as well: their dwell is a lower bound
    (real enter time unknown), but in production where detection is started
    once, these are a small minority and folding them in keeps the stats
    in line with intuition. The ``session_start_count`` column tells the
    user how many of the rows are partial measurements. ``lost_id`` enters
    are still excluded — their exit was missed, so the dwell is unreliable
    in a way no flag can correct."""
    by_zone: dict[str, dict] = {}
    for i, ev in enumerate(buffer):
        et = ev.get("event_type", "")
        if et not in ZONE_PER_OBJECT_TYPES:
            continue
        name = ev.get("region_name", "") or "(unnamed)"
        b = by_zone.setdefault(name, {
            "enter_total": 0, "enter_clean": 0, "exit_count": 0,
            "session_start": 0, "still_inside": 0, "lost_id": 0,
            "dwells": [], "longest": (None, -1.0, False, False),
            "objs": set(),
        })
        b["objs"].add(int(ev.get("object_id", 0)))
        pairing_info = pairings.get(i, {"pairing": "n/a", "dwell_sec": None})
        pairing = pairing_info.get("pairing", "n/a")
        if et == "zone_enter":
            b["enter_total"] += 1
            quality = _ev_quality(ev)
            if pairing == "matched" and quality == "ok":
                b["enter_clean"] += 1
            # Dwell goes into stats for matched + still_inside, regardless
            # of quality. session_start dwells are lower bounds but kept in
            # so the small handful of "already inside at start" cars don't
            # silently disappear from the average.
            if pairing in ("matched", "still_inside"):
                d = pairing_info.get("dwell_sec")
                if d is not None:
                    b["dwells"].append(d)
                    if d > b["longest"][1]:
                        b["longest"] = (
                            f"{ev.get('class_name','')}#{ev.get('object_id',0)}",
                            d,
                            pairing == "still_inside",
                            quality == "session_start")
            if pairing == "still_inside":
                b["still_inside"] += 1
            elif pairing == "lost_id":
                b["lost_id"] += 1
            if quality == "session_start":
                b["session_start"] += 1
        elif et == "zone_exit":
            b["exit_count"] += 1

    rows: list[dict] = []
    for name, b in sorted(by_zone.items()):
        dwells = b["dwells"]
        avg = sum(dwells) / len(dwells) if dwells else None
        mn = min(dwells) if dwells else None
        mx = max(dwells) if dwells else None
        longest_obj = ""
        if b["longest"][1] >= 0:
            longest_obj = b["longest"][0]
            markers: list[str] = []
            if b["longest"][2]:
                markers.append("still inside")
            if b["longest"][3]:
                markers.append("session start")
            if markers:
                longest_obj += " (" + ", ".join(markers) + ")"
        rows.append({
            "type": "zone", "name": name,
            "enter_count_total": b["enter_total"],
            "enter_count_clean": b["enter_clean"],
            "exit_count": b["exit_count"],
            "avg_dwell_sec": avg, "min_dwell_sec": mn, "max_dwell_sec": mx,
            "longest_object": longest_obj,
            "session_start_count": b["session_start"],
            "still_inside_count": b["still_inside"],
            "lost_id_count": b["lost_id"],
            "unique_objects": len(b["objs"]),
        })
    return rows


def compute_area_pairings(buffer: list[dict],
                          entrance_ids: set[str],
                          exit_ids: set[str],
                          currently_tracked_ids: set[int],
                          export_time: float,
                          bidirectional_ids: set[str] | None = None) -> list[dict]:
    """Compute per-visit pathway dwell — time spent in the watched area
    *outside* of any zone (i.e., on the road, not in a parking spot).

    Conceptual model:
      - Crossing the entrance line puts the object **in area** (on the
        pathway).
      - Entering a zone takes them **out of area** (now in a zone).
      - Exiting a zone puts them **back in area** (back on the pathway).
      - Crossing the exit line ends the visit entirely.

    Each "visit" is one entrance→exit cycle. Within a visit, area_dwell
    accumulates only while ``in_area AND NOT in_zone``. Multiple
    pathway segments per visit get summed into a single record.

    Returns one record per visit::

        {"object_id", "class_name", "dwell_sec", "pairing"}

    pairing values: ``matched`` (visit ended via the exit line or a
    bidirectional line reverse-direction crossing), ``reversed`` (visit
    ended via wrong-direction crossing on a strict entrance/exit line —
    real reverse driving and camera-shake artifacts both land here, since
    they look identical in the data), ``still_inside`` (visit unfinished,
    object still tracked), ``lost_id`` (visit unfinished, object no
    longer tracked).
    """
    # Collect all events that affect area/zone state, per object.
    by_obj: dict[int, list[tuple]] = {}
    for ev in buffer:
        et = ev.get("event_type")
        if et not in ("line_in", "line_out", "zone_enter", "zone_exit"):
            continue
        rid = str(ev.get("region_id", ""))
        oid = int(ev.get("object_id", 0))
        ts = float(ev.get("timestamp", 0.0))
        cls = ev.get("class_name", "")
        is_entrance = rid in entrance_ids
        is_exit = rid in exit_ids
        by_obj.setdefault(oid, []).append(
            (ts, et, rid, is_entrance, is_exit, cls))

    bidir = set(bidirectional_ids or [])

    pairs: list[dict] = []
    for oid, events in by_obj.items():
        events.sort(key=lambda e: e[0])
        in_area = False
        in_zone = False
        # Start of the current pathway segment (in_area AND NOT in_zone).
        # None when the object is not currently accruing pathway time.
        seg_start: float | None = None
        # Total pathway time accumulated for the in-progress visit.
        visit_total = 0.0
        visit_started = False
        visit_open_ts: float | None = None
        # Per-segment intervals collected so hourly aggregation can clip
        # them to hour boundaries. Each tuple is (seg_start, seg_end).
        visit_segments: list[tuple[float, float]] = []
        last_class = ""

        def _close_segment(now_ts: float) -> None:
            nonlocal seg_start, visit_total
            if seg_start is not None:
                end = max(seg_start, now_ts)
                visit_total += end - seg_start
                visit_segments.append((seg_start, end))
                seg_start = None

        def _emit_visit(pairing: str, end_ts: float) -> None:
            nonlocal visit_total, visit_started, visit_open_ts
            pairs.append({
                "object_id": oid, "class_name": last_class,
                "dwell_sec": visit_total, "pairing": pairing,
                "visit_start_ts": visit_open_ts or 0.0,
                "visit_end_ts": end_ts,
                "segments": list(visit_segments)})
            visit_total = 0.0
            visit_started = False
            visit_open_ts = None
            visit_segments.clear()

        for ts, et, rid, is_ent, is_ext, cls in events:
            if cls:
                last_class = cls

            # Entrance crossing → start a fresh visit on the pathway.
            if et == "line_in" and is_ent:
                if visit_started:
                    # Defensive: previous visit never closed; treat its
                    # accumulated time as a still_inside-style record.
                    _close_segment(ts)
                    _emit_visit("still_inside", ts)
                in_area = True
                in_zone = False
                seg_start = ts
                visit_total = 0.0
                visit_started = True
                visit_open_ts = ts
                visit_segments.clear()

            # Exit-line crossing — visit ends. Two cases:
            # - line_in on a line whose function is exit (forward direction)
            # - line_out on a *bidirectional* line (reverse direction on a
            #   single-line boundary = leaving area).
            elif visit_started and (
                    (et == "line_in" and is_ext)
                    or (et == "line_out" and rid in bidir)):
                _close_segment(ts)
                _emit_visit("matched", ts)
                in_area = False
                in_zone = False

            # Reverse crossing on a strict line — wrong-direction crossing
            # while the visit was open. Covers both real reverse-driving
            # cars and camera-shake artifacts; treated identically because
            # they look the same in the data, and shake is a deployment
            # issue (camera mounting), not a tracker correctness issue.
            # Recorded so the data isn't silently dropped, but excluded
            # from dwell aggregates so it doesn't pollute avg/min/max.
            elif visit_started and et == "line_out" and (is_ent or is_ext) \
                    and rid not in bidir:
                _close_segment(ts)
                _emit_visit("reversed", ts)
                in_area = False
                in_zone = False

            # Zone enter — leave pathway, enter zone (still in area
            # bounding box but no longer accruing pathway time).
            elif et == "zone_enter" and visit_started and not in_zone:
                _close_segment(ts)
                in_zone = True

            # Zone exit — back on pathway, accrue again from now.
            elif et == "zone_exit" and visit_started and in_zone:
                in_zone = False
                seg_start = ts

        # End of buffer — visit unfinished.
        if visit_started:
            if not in_zone:
                _close_segment(export_time)
            pairing = ("still_inside" if oid in currently_tracked_ids
                       else "lost_id")
            _emit_visit(pairing, export_time)

    return pairs


def _summary_for_area(buffer: list[dict],
                      area_pairings: list[dict]) -> list[dict]:
    """Area-level row in summary.csv. Dwell stats come from line crossings
    (entrance→exit pairings), NOT from zone enters — those would just
    duplicate the per-zone numbers. ``stuck_event_count`` still reflects
    the number of group "stuck" events that fired.

    Zero-dwell visits are filtered from avg/min/max because they're
    tracker artifacts from simultaneous events (entrance and exit firing
    in the same frame, an ID switch teleporting through both lines, etc.).
    A real pathway crossing always takes physical time. The filtered count
    is surfaced via ``anomaly_count`` so it isn't silently dropped.

    Reverse-direction crossings on strict lines (camera shake or actual
    reverse driving — same data shape, treated identically) close the
    visit as ``pairing="reversed"``. Counted via ``reversed_count`` but
    excluded from dwell aggregates so they don't pull avg/min/max around.
    """
    stuck_rows = [ev for ev in buffer
                  if ev.get("event_type") in AREA_EVENT_TYPES]
    # Skip area row entirely if there's no area data at all (no entrance/exit
    # crossings AND no stuck events). Otherwise the row would be all blank.
    if not area_pairings and not stuck_rows:
        return []

    # Dwell aggregates use matched + still_inside only. Excluded:
    # - lost_id (actual exit time unknown)
    # - reversed (wrong-direction crossing — physically completed, but
    #   the timing is contaminated by camera shake / reverse driving and
    #   shouldn't pull avg/min/max around)
    # Plus a positive-dwell filter to drop same-frame artifacts.
    dwells: list[float] = []
    longest = (None, -1.0, False)
    objs: set[int] = set()
    still_inside = lost_id = anomaly = reversed_ct = 0
    for p in area_pairings:
        objs.add(p["object_id"])
        pairing = p["pairing"]
        d = p["dwell_sec"]
        if pairing == "still_inside":
            still_inside += 1
        elif pairing == "lost_id":
            lost_id += 1
        elif pairing == "reversed":
            reversed_ct += 1
        if pairing in ("matched", "still_inside"):
            if d <= 0.0:
                anomaly += 1
                continue
            dwells.append(d)
            if d > longest[1]:
                longest = (
                    f"{p['class_name']}#{p['object_id']}",
                    d, pairing == "still_inside")
    avg = sum(dwells) / len(dwells) if dwells else None
    longest_obj = ""
    if longest[1] >= 0:
        longest_obj = longest[0]
        if longest[2]:
            longest_obj += " (still inside)"

    matched_count = sum(1 for p in area_pairings if p["pairing"] == "matched")
    return [{
        "type": "area", "name": "Area",
        "enter_count_total": len(area_pairings),
        "enter_count_clean": matched_count,
        "exit_count": matched_count,
        "stuck_event_count": len(stuck_rows),
        "avg_dwell_sec": avg,
        "min_dwell_sec": min(dwells) if dwells else None,
        "max_dwell_sec": max(dwells) if dwells else None,
        "longest_object": longest_obj,
        "still_inside_count": still_inside,
        "lost_id_count": lost_id,
        "anomaly_count": anomaly,
        "reversed_count": reversed_ct,
        "unique_objects": len(objs),
    }]


SUMMARY_HEADERS = [
    "type", "name",
    "in_count", "out_count",
    "enter_count_total", "enter_count_clean", "exit_count",
    "avg_dwell_sec", "min_dwell_sec", "max_dwell_sec", "longest_object",
    "session_start_count", "still_inside_count", "lost_id_count",
    "anomaly_count", "reversed_count",
    "stuck_event_count", "unique_objects",
]


def build_summary_rows(buffer: list[dict],
                       pairings: dict[int, dict],
                       area_pairings: list[dict] | None = None) -> list[list]:
    area_pairings = area_pairings or []
    rows: list[list] = []
    for r in (_summary_for_lines(buffer)
              + _summary_for_zones(buffer, pairings)
              + _summary_for_area(buffer, area_pairings)):
        # Area dwell uses 2 decimals so very fast crossings (~0.04s @ 25fps)
        # don't get misleadingly rounded down to 0.0. Other rows use the
        # standard 1-decimal format.
        decimals = 2 if r.get("type") == "area" else 1
        rows.append([
            r.get("type", ""),
            r.get("name", ""),
            r.get("in_count", ""),
            r.get("out_count", ""),
            r.get("enter_count_total", ""),
            r.get("enter_count_clean", ""),
            r.get("exit_count", ""),
            _format_dwell(r.get("avg_dwell_sec"), decimals),
            _format_dwell(r.get("min_dwell_sec"), decimals),
            _format_dwell(r.get("max_dwell_sec"), decimals),
            r.get("longest_object", ""),
            r.get("session_start_count", ""),
            r.get("still_inside_count", ""),
            r.get("lost_id_count", ""),
            r.get("anomaly_count", ""),
            r.get("reversed_count", ""),
            r.get("stuck_event_count", ""),
            r.get("unique_objects", ""),
        ])
    return rows


# ---------------------------------------------------------------------------
# Hourly + Daily summary
# ---------------------------------------------------------------------------
#
# Hourly_Summary slices the buffer into top-of-hour buckets (local time) and
# emits the same row shape as the whole-session Summary, plus leading
# ``date`` and ``hour`` columns. Dwell stats inside each hour are *clipped*
# to the hour boundaries — a 2.5-hour visit that crosses three hour
# boundaries shows ~5/60/60/25 minutes of clipped dwell across four rows,
# not its full 150 minutes in any one row.
#
# Daily_Summary then ranks the hourly rows per calendar day and surfaces
# the peaks: which hour had the most entrances, the longest *full* visit
# (not the clipped portion — the real entrance-to-exit duration), etc.
# That keeps the "object stayed 2:30 across hours" intuition intact even
# though the per-hour cells split it up.

HOURLY_HEADERS = ["date", "hour"] + SUMMARY_HEADERS

DAILY_HEADERS = ["date", "metric", "hour", "value", "object"]


def _bucket_label(hour_start_ts: float) -> tuple[str, str]:
    """``(date_str, hour_label)`` for the hour bucket starting at ``ts``."""
    lt = time.localtime(hour_start_ts)
    date_str = time.strftime("%Y-%m-%d", lt)
    hour_label = (f"{lt.tm_hour:02d}:00 - "
                  f"{(lt.tm_hour + 1) % 24:02d}:00")
    return date_str, hour_label


def _iter_hour_buckets(start_ts: float,
                       end_ts: float) -> list[tuple[float, float]]:
    """Yield ``(hour_start_ts, hour_end_ts)`` for every hour boundary touched
    by the [start_ts, end_ts] range (local time, top-of-hour aligned).
    Empty buckets are included so the report has a visible row for every
    hour of the run."""
    if end_ts <= start_ts:
        return []
    lt = time.localtime(start_ts)
    # Floor to the start of the hour in local time.
    floor_struct = time.struct_time((lt.tm_year, lt.tm_mon, lt.tm_mday,
                                     lt.tm_hour, 0, 0,
                                     lt.tm_wday, lt.tm_yday, lt.tm_isdst))
    h0 = time.mktime(floor_struct)
    buckets: list[tuple[float, float]] = []
    h = h0
    # 3700s instead of 3600 absorbs DST forward jumps without skipping; the
    # next bucket realigns via mktime on the new local-hour struct.
    while h < end_ts:
        # Walk to the next top-of-hour by adding ~1h then re-floored.
        next_lt = time.localtime(h + 3700)
        next_floor = time.struct_time((next_lt.tm_year, next_lt.tm_mon,
                                       next_lt.tm_mday, next_lt.tm_hour,
                                       0, 0, next_lt.tm_wday,
                                       next_lt.tm_yday, next_lt.tm_isdst))
        h_next = time.mktime(next_floor)
        if h_next <= h:
            h_next = h + 3600  # last-resort guard against pathological tz
        buckets.append((h, h_next))
        h = h_next
    return buckets


def _clip(seg_start: float, seg_end: float,
          h_start: float, h_end: float) -> float:
    """Overlap of [seg_start, seg_end] with [h_start, h_end] in seconds."""
    s = max(seg_start, h_start)
    e = min(seg_end, h_end)
    return e - s if e > s else 0.0


def compute_zone_visits(buffer: list[dict],
                        pairings: dict[int, dict],
                        export_time: float) -> list[dict]:
    """Per-visit records for zones, with full enter/exit timestamps.

    Returns a list of dicts::

        {"zone_name", "object_id", "class_name",
         "enter_ts", "exit_ts", "dwell_sec",
         "pairing", "quality"}

    ``exit_ts`` is the actual zone_exit time for ``matched`` visits, and
    ``export_time`` for ``still_inside`` / ``lost_id``. Used by the hourly
    aggregator to clip dwell to hour boundaries.
    """
    visits: list[dict] = []
    for i, ev in enumerate(buffer):
        if ev.get("event_type") != "zone_enter":
            continue
        info = pairings.get(i)
        if not info:
            continue
        pairing = info.get("pairing", "n/a")
        if pairing == "n/a":
            continue
        enter_ts = float(ev.get("timestamp", 0.0))
        dwell = info.get("dwell_sec")
        if dwell is None:
            continue
        exit_ts = enter_ts + float(dwell)
        visits.append({
            "zone_name": ev.get("region_name") or "(unnamed)",
            "object_id": int(ev.get("object_id", 0)),
            "class_name": ev.get("class_name", ""),
            "enter_ts": enter_ts,
            "exit_ts": exit_ts,
            "dwell_sec": float(dwell),
            "pairing": pairing,
            "quality": _ev_quality(ev),
        })
    return visits


def _hourly_lines_block(buffer: list[dict], line_names: list[str],
                        h_start: float, h_end: float) -> list[dict]:
    by_line: dict[str, dict] = {
        name: {"in": 0, "out": 0, "objs": set()} for name in line_names}
    for ev in buffer:
        et = ev.get("event_type")
        if et not in LINE_EVENT_TYPES:
            continue
        ts = float(ev.get("timestamp", 0.0))
        if not (h_start <= ts < h_end):
            continue
        name = ev.get("region_name") or "(unnamed)"
        b = by_line.setdefault(name, {"in": 0, "out": 0, "objs": set()})
        if et in ("line_in", "line_cross_in"):
            b["in"] += 1
        else:
            b["out"] += 1
        b["objs"].add(int(ev.get("object_id", 0)))
    return [{
        "type": "line", "name": n,
        "in_count": by_line[n]["in"],
        "out_count": by_line[n]["out"],
        "unique_objects": len(by_line[n]["objs"]),
    } for n in sorted(by_line.keys())]


def _hourly_zones_block(buffer: list[dict],
                        pairings: dict[int, dict],
                        zone_visits: list[dict],
                        zone_names: list[str],
                        h_start: float, h_end: float) -> list[dict]:
    by_zone: dict[str, dict] = {n: {
        "enter_total": 0, "enter_clean": 0, "exit_count": 0,
        "session_start": 0, "still_inside": 0, "lost_id": 0,
        "dwells": [],  # list[(clipped_sec, obj_label, quality)]
        "objs": set(),
    } for n in zone_names}

    # Event-time counts (enters/exits/session_start) — by event timestamp.
    for i, ev in enumerate(buffer):
        et = ev.get("event_type")
        if et not in ZONE_PER_OBJECT_TYPES:
            continue
        ts = float(ev.get("timestamp", 0.0))
        if not (h_start <= ts < h_end):
            continue
        name = ev.get("region_name") or "(unnamed)"
        b = by_zone.setdefault(name, {
            "enter_total": 0, "enter_clean": 0, "exit_count": 0,
            "session_start": 0, "still_inside": 0, "lost_id": 0,
            "dwells": [], "objs": set(),
        })
        b["objs"].add(int(ev.get("object_id", 0)))
        if et == "zone_enter":
            b["enter_total"] += 1
            quality = _ev_quality(ev)
            pairing = pairings.get(i, {}).get("pairing", "n/a")
            if pairing == "matched" and quality == "ok":
                b["enter_clean"] += 1
            if quality == "session_start":
                b["session_start"] += 1
            if pairing == "lost_id":
                b["lost_id"] += 1
        elif et == "zone_exit":
            b["exit_count"] += 1

    # Clipped-dwell + carryover (still_inside-at-end-of-hour) — by visit.
    for v in zone_visits:
        if v["pairing"] == "lost_id":
            continue  # excluded from dwell — exit time unreliable
        clipped = _clip(v["enter_ts"], v["exit_ts"], h_start, h_end)
        if clipped <= 0:
            continue
        b = by_zone.setdefault(v["zone_name"], {
            "enter_total": 0, "enter_clean": 0, "exit_count": 0,
            "session_start": 0, "still_inside": 0, "lost_id": 0,
            "dwells": [], "objs": set(),
        })
        b["objs"].add(v["object_id"])
        b["dwells"].append((
            clipped,
            f"{v['class_name']}#{v['object_id']}",
            v["quality"]))
        # Carryover marker: visit still in zone at end of this hour.
        if v["enter_ts"] < h_end and v["exit_ts"] > h_end:
            b["still_inside"] += 1

    rows: list[dict] = []
    for n in sorted(by_zone.keys()):
        b = by_zone[n]
        dwells = [d[0] for d in b["dwells"]]
        avg = sum(dwells) / len(dwells) if dwells else None
        mn = min(dwells) if dwells else None
        mx = max(dwells) if dwells else None
        longest_obj = ""
        if b["dwells"]:
            longest = max(b["dwells"], key=lambda d: d[0])
            longest_obj = longest[1]
            if longest[2] == "session_start":
                longest_obj += " (session start)"
        rows.append({
            "type": "zone", "name": n,
            "enter_count_total": b["enter_total"],
            "enter_count_clean": b["enter_clean"],
            "exit_count": b["exit_count"],
            "avg_dwell_sec": avg,
            "min_dwell_sec": mn,
            "max_dwell_sec": mx,
            "longest_object": longest_obj,
            "session_start_count": b["session_start"],
            "still_inside_count": b["still_inside"],
            "lost_id_count": b["lost_id"],
            "unique_objects": len(b["objs"]),
        })
    return rows


def _hourly_area_block(buffer: list[dict],
                       area_pairings: list[dict],
                       entrance_ids: set[str], exit_ids: set[str],
                       bidirectional_ids: set[str],
                       h_start: float, h_end: float) -> list[dict]:
    in_count = out_count = stuck = 0
    objs: set[int] = set()

    for ev in buffer:
        ts = float(ev.get("timestamp", 0.0))
        if not (h_start <= ts < h_end):
            continue
        et = ev.get("event_type")
        rid = str(ev.get("region_id", ""))
        oid = int(ev.get("object_id", 0))
        if et in AREA_EVENT_TYPES:
            stuck += 1
            objs.add(oid)
            continue
        if et in ("line_in", "line_out") and (
                rid in entrance_ids or rid in exit_ids
                or rid in bidirectional_ids):
            objs.add(oid)
            if et in ("line_in", "line_cross_in"):
                in_count += 1
            else:
                out_count += 1

    # Clipped pathway dwell — sum per visit, longest-per-hour by visit.
    dwells: list[tuple[float, str]] = []
    still_inside_at_end = lost_id_in_hour = reversed_in_hour = anomaly = 0
    matched_in_hour = 0
    for p in area_pairings:
        segs = p.get("segments") or []
        clipped_total = 0.0
        for s, e in segs:
            clipped_total += _clip(s, e, h_start, h_end)
        if clipped_total > 0:
            label = f"{p.get('class_name','')}#{p['object_id']}"
            dwells.append((clipped_total, label))
            objs.add(p["object_id"])
        v_start = p.get("visit_start_ts", 0.0)
        v_end = p.get("visit_end_ts", 0.0)
        # Carryover: pathway visit still open at end of this hour.
        if (p["pairing"] in ("matched", "still_inside", "reversed", "lost_id")
                and v_start < h_end and v_end > h_end):
            still_inside_at_end += 1
        # End-of-visit attribution (lost_id / reversed / matched buckets
        # follow the *visit_end* event time, which is the line crossing or
        # the export boundary).
        if h_start <= v_end < h_end:
            if p["pairing"] == "lost_id":
                lost_id_in_hour += 1
            elif p["pairing"] == "reversed":
                reversed_in_hour += 1
            elif p["pairing"] == "matched":
                matched_in_hour += 1
                if p.get("dwell_sec", 0.0) <= 0.0:
                    anomaly += 1

    # Skip the area row entirely if this hour has no area activity at all —
    # avoids a pure-zero row in hours where only zone activity happened.
    if (in_count == 0 and out_count == 0 and stuck == 0 and not dwells
            and still_inside_at_end == 0 and lost_id_in_hour == 0
            and reversed_in_hour == 0 and matched_in_hour == 0):
        return []

    dwell_values = [d[0] for d in dwells]
    avg = sum(dwell_values) / len(dwell_values) if dwell_values else None
    longest_obj = ""
    if dwells:
        longest_obj = max(dwells, key=lambda d: d[0])[1]
    return [{
        "type": "area", "name": "Area",
        "in_count": in_count,
        "out_count": out_count,
        "enter_count_total": (matched_in_hour + still_inside_at_end
                              + lost_id_in_hour + reversed_in_hour),
        "enter_count_clean": matched_in_hour,
        "exit_count": matched_in_hour,
        "avg_dwell_sec": avg,
        "min_dwell_sec": min(dwell_values) if dwell_values else None,
        "max_dwell_sec": max(dwell_values) if dwell_values else None,
        "longest_object": longest_obj,
        "still_inside_count": still_inside_at_end,
        "lost_id_count": lost_id_in_hour,
        "anomaly_count": anomaly,
        "reversed_count": reversed_in_hour,
        "stuck_event_count": stuck,
        "unique_objects": len(objs),
    }]


def _all_region_names(buffer: list[dict]) -> tuple[list[str], list[str]]:
    """Distinct line names and zone names appearing anywhere in the buffer.
    Used to keep empty hour buckets visible — every line and zone gets a
    row in every hour, even if all values are zero."""
    lines: set[str] = set()
    zones: set[str] = set()
    for ev in buffer:
        et = ev.get("event_type")
        name = ev.get("region_name") or "(unnamed)"
        if et in LINE_EVENT_TYPES:
            lines.add(name)
        elif et in ZONE_PER_OBJECT_TYPES:
            zones.add(name)
    return sorted(lines), sorted(zones)


def _row_to_summary_cells(r: dict, decimals: int) -> list:
    """Map a per-region dict (lines/zones/area shape) to a SUMMARY_HEADERS
    row, with empty strings for cells that don't apply to this row type."""
    return [
        r.get("type", ""),
        r.get("name", ""),
        r.get("in_count", ""),
        r.get("out_count", ""),
        r.get("enter_count_total", ""),
        r.get("enter_count_clean", ""),
        r.get("exit_count", ""),
        _format_dwell(r.get("avg_dwell_sec"), decimals),
        _format_dwell(r.get("min_dwell_sec"), decimals),
        _format_dwell(r.get("max_dwell_sec"), decimals),
        r.get("longest_object", ""),
        r.get("session_start_count", ""),
        r.get("still_inside_count", ""),
        r.get("lost_id_count", ""),
        r.get("anomaly_count", ""),
        r.get("reversed_count", ""),
        r.get("stuck_event_count", ""),
        r.get("unique_objects", ""),
    ]


def build_hourly_rows(buffer: list[dict],
                      pairings: dict[int, dict],
                      area_pairings: list[dict],
                      metadata: dict,
                      buffer_started_at: float,
                      export_time: float) -> list[list]:
    """Per-hour blocks. Each block: line rows, then zone rows, then area row
    (when applicable). Blocks separated by a blank row. Returns rows shaped
    for HOURLY_HEADERS (date, hour, ...summary columns)."""
    line_names, zone_names = _all_region_names(buffer)
    zone_visits = compute_zone_visits(buffer, pairings, export_time)
    entrance_ids = set(metadata.get("entrance_line_ids") or [])
    exit_ids = set(metadata.get("exit_line_ids") or [])
    bidirectional_ids = set(metadata.get("bidirectional_line_ids") or [])

    out: list[list] = []
    buckets = _iter_hour_buckets(buffer_started_at, export_time)
    leading_pad = ["", ""]  # (date, hour) — empty for blank separator
    for idx, (h_start, h_end) in enumerate(buckets):
        date_str, hour_label = _bucket_label(h_start)
        line_rows = _hourly_lines_block(buffer, line_names, h_start, h_end)
        zone_rows = _hourly_zones_block(
            buffer, pairings, zone_visits, zone_names, h_start, h_end)
        area_rows = _hourly_area_block(
            buffer, area_pairings, entrance_ids, exit_ids,
            bidirectional_ids, h_start, h_end)
        # Attach date/hour prefix to each row.
        for r in line_rows:
            out.append([date_str, hour_label]
                       + _row_to_summary_cells(r, decimals=1))
        for r in zone_rows:
            out.append([date_str, hour_label]
                       + _row_to_summary_cells(r, decimals=1))
        for r in area_rows:
            out.append([date_str, hour_label]
                       + _row_to_summary_cells(r, decimals=2))
        # Blank separator row between hour blocks (not after the last one).
        if idx < len(buckets) - 1:
            out.append(leading_pad + [""] * len(SUMMARY_HEADERS))
    return out


def build_daily_rows(buffer: list[dict],
                     pairings: dict[int, dict],
                     area_pairings: list[dict],
                     metadata: dict,
                     buffer_started_at: float,
                     export_time: float) -> list[list]:
    """Per-day rankings. For each calendar day in the run, surface the hour
    that "won" each metric. Dwell metrics use the *full* visit dwell (not
    the hour-clipped portion), so a 2.5-hour visit shows its real duration
    and is attributed to the hour the visit ended in."""
    line_names, zone_names = _all_region_names(buffer)
    zone_visits = compute_zone_visits(buffer, pairings, export_time)
    entrance_ids = set(metadata.get("entrance_line_ids") or [])
    exit_ids = set(metadata.get("exit_line_ids") or [])
    bidirectional_ids = set(metadata.get("bidirectional_line_ids") or [])

    # Per-hour aggregates (date, hour_label, hour_start) keyed for ranking.
    buckets = _iter_hour_buckets(buffer_started_at, export_time)
    per_hour: list[dict] = []
    for h_start, h_end in buckets:
        date_str, hour_label = _bucket_label(h_start)
        line_rows = _hourly_lines_block(buffer, line_names, h_start, h_end)
        zone_rows = _hourly_zones_block(
            buffer, pairings, zone_visits, zone_names, h_start, h_end)
        area_rows = _hourly_area_block(
            buffer, area_pairings, entrance_ids, exit_ids,
            bidirectional_ids, h_start, h_end)
        per_hour.append({
            "date": date_str, "hour": hour_label,
            "h_start": h_start, "h_end": h_end,
            "entrance_in": sum(r["in_count"] for r in line_rows),
            "exit_out": sum(r["out_count"] for r in line_rows),
            "zone_enters": sum(r.get("enter_count_total", 0)
                               for r in zone_rows),
            "zone_exits": sum(r.get("exit_count", 0) for r in zone_rows),
            "stuck": sum(r.get("stuck_event_count", 0) for r in area_rows),
            "reversed": sum(r.get("reversed_count", 0) for r in area_rows),
            "active_objs": (
                sum(r.get("unique_objects", 0) for r in line_rows)
                + sum(r.get("unique_objects", 0) for r in zone_rows)),
        })

    # Group hours by date so each calendar day gets its own ranking block.
    by_date: dict[str, list[dict]] = {}
    for h in per_hour:
        by_date.setdefault(h["date"], []).append(h)

    # Full-dwell visits with their *ending* hour, used for longest/shortest
    # rankings. Lost_id visits are excluded — exit time is unreliable.
    completed_zone = [v for v in zone_visits if v["pairing"] == "matched"]
    completed_area = [p for p in area_pairings
                      if p["pairing"] == "matched" and p.get("dwell_sec", 0) > 0]

    def _hour_label_for(ts: float) -> tuple[str, str]:
        lt = time.localtime(ts)
        floor = time.mktime(time.struct_time(
            (lt.tm_year, lt.tm_mon, lt.tm_mday, lt.tm_hour, 0, 0,
             lt.tm_wday, lt.tm_yday, lt.tm_isdst)))
        return _bucket_label(floor)

    rows: list[list] = []
    for date_str in sorted(by_date.keys()):
        hours = by_date[date_str]

        def _peak(metric_key: str) -> tuple[str, int]:
            best = max(hours, key=lambda h: h[metric_key], default=None)
            if best is None or best[metric_key] == 0:
                return "", 0
            return best["hour"], best[metric_key]

        for label, key in [
            ("Peak entrance line crossings", "entrance_in"),
            ("Peak exit line crossings", "exit_out"),
            ("Peak zone enters", "zone_enters"),
            ("Peak zone exits", "zone_exits"),
            ("Most stuck events", "stuck"),
            ("Most reversed crossings", "reversed"),
            ("Most active objects", "active_objs"),
        ]:
            hour_lbl, val = _peak(key)
            rows.append([date_str, label, hour_lbl, val if val else "", ""])

        # Longest / shortest dwell — full visit dwell, attributed to the
        # hour the visit ended in. Filtered to visits ending on this date.
        zones_today = [v for v in completed_zone
                       if _hour_label_for(v["exit_ts"])[0] == date_str]
        if zones_today:
            longest_z = max(zones_today, key=lambda v: v["dwell_sec"])
            shortest_z = min(zones_today, key=lambda v: v["dwell_sec"])
            rows.append([
                date_str, "Longest zone visit (completed)",
                _hour_label_for(longest_z["exit_ts"])[1],
                _format_dwell(longest_z["dwell_sec"], 1),
                f"{longest_z['class_name']}#{longest_z['object_id']} "
                f"({longest_z['zone_name']})",
            ])
            rows.append([
                date_str, "Shortest zone visit (completed)",
                _hour_label_for(shortest_z["exit_ts"])[1],
                _format_dwell(shortest_z["dwell_sec"], 1),
                f"{shortest_z['class_name']}#{shortest_z['object_id']} "
                f"({shortest_z['zone_name']})",
            ])
        else:
            rows.append([date_str, "Longest zone visit (completed)",
                         "", "", ""])
            rows.append([date_str, "Shortest zone visit (completed)",
                         "", "", ""])

        area_today = [p for p in completed_area
                      if _hour_label_for(p["visit_end_ts"])[0] == date_str]
        if area_today:
            longest_a = max(area_today, key=lambda p: p["dwell_sec"])
            shortest_a = min(area_today, key=lambda p: p["dwell_sec"])
            rows.append([
                date_str, "Longest area visit (completed)",
                _hour_label_for(longest_a["visit_end_ts"])[1],
                _format_dwell(longest_a["dwell_sec"], 2),
                f"{longest_a.get('class_name','')}#{longest_a['object_id']}",
            ])
            rows.append([
                date_str, "Shortest area visit (completed)",
                _hour_label_for(shortest_a["visit_end_ts"])[1],
                _format_dwell(shortest_a["dwell_sec"], 2),
                f"{shortest_a.get('class_name','')}#{shortest_a['object_id']}",
            ])
        else:
            rows.append([date_str, "Longest area visit (completed)",
                         "", "", ""])
            rows.append([date_str, "Shortest area visit (completed)",
                         "", "", ""])

        # Blank separator between dates.
        if date_str != sorted(by_date.keys())[-1]:
            rows.append(["", "", "", "", ""])
    return rows


def build_metadata_kv(metadata: dict, buffer: list[dict],
                      pairings: dict[int, dict]) -> list[list]:
    """Key/value rows for the XLSX Metadata sheet. Surfaces export context
    (project, model, source, mode mix) plus a session-level quality
    summary (lost_id rate, session_start count) so an exported report is
    self-describing without needing a sidecar file."""
    line_count = sum(1 for ev in buffer
                     if ev.get("event_type") in LINE_EVENT_TYPES)
    zone_count = sum(1 for ev in buffer
                     if ev.get("event_type") in ZONE_PER_OBJECT_TYPES)
    area_count = sum(1 for ev in buffer
                     if ev.get("event_type") in AREA_EVENT_TYPES)
    test_count = sum(1 for ev in buffer if _ev_mode(ev) == "TEST")
    live_count = sum(1 for ev in buffer if _ev_mode(ev) == "LIVE")
    enter_total = sum(1 for ev in buffer if ev.get("event_type") == "zone_enter")
    lost_id_total = sum(1 for info in pairings.values()
                        if info.get("pairing") == "lost_id")
    still_inside_total = sum(1 for info in pairings.values()
                             if info.get("pairing") == "still_inside")
    session_start_total = sum(1 for ev in buffer
                              if ev.get("event_type") == "zone_enter"
                              and _ev_quality(ev) == "session_start")
    lost_rate = (lost_id_total / enter_total) if enter_total else 0.0

    src_str = (f"{metadata.get('source_type', '')} | "
               f"{metadata.get('source_value', '')} | "
               f"{metadata.get('source_resolution', '')}")
    if metadata.get("source_fps"):
        src_str += f" @ {metadata.get('source_fps', 0):.1f} fps"

    rows = [
        ["Export created", _iso(metadata.get("export_time", time.time()))],
        ["Project", metadata.get("project_name", "")],
        ["Project file", metadata.get("project_path", "")],
        ["Model", metadata.get("model_name", "")],
        ["Source", src_str],
        ["Inference size", str(metadata.get("imgsz", ""))],
        ["Inference scale", metadata.get("scale_str", "")],
        ["", ""],
        ["Buffer started", _iso(metadata.get("buffer_started_at", 0))],
        ["Buffer ended", _iso(metadata.get("export_time", time.time()))
         + ("  (snapshot at export time, detection still running)"
            if metadata.get("still_running") else "")],
        ["Total events", len(buffer)],
        ["   Lines", line_count],
        ["   Zones (object)", zone_count],
        ["   Area (stuck)", area_count],
        ["", ""],
        ["Mode mix", f"LIVE={live_count}, TEST={test_count}"],
        ["", ""],
        ["Quality summary", ""],
        ["   total_enters", enter_total],
        ["   session_start_enters", session_start_total],
        ["   still_inside_at_export", still_inside_total],
        ["   lost_id_count", lost_id_total],
        ["   lost_id_rate", f"{lost_rate * 100:.1f}%"],
    ]
    if lost_rate > 0.05:
        rows.append([
            "   tracker_quality_hint",
            "lost_id_rate above 5% — ID switches likely affecting dwell measurements"])
    return rows


def write_xlsx(folder: str, buffer: list[dict],
               pairings: dict[int, dict], metadata: dict,
               area_pairings: list[dict] | None = None) -> str | None:
    """Write the multi-sheet ``report.xlsx`` workbook. Sole output of an
    export. Returns the file path on success, or None if openpyxl isn't
    installed (caller should surface that to the user).

    Sheets, in order: Lines, Zones, Summary, Hourly_Summary, Daily_Summary,
    Metadata.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
    except ImportError:
        return None

    area_pairings = area_pairings or []
    buffer_started_at = float(metadata.get("buffer_started_at", 0.0))
    export_time = float(metadata.get("export_time", time.time()))
    if buffer_started_at <= 0 and buffer:
        buffer_started_at = float(buffer[0].get("timestamp", export_time))

    wb = Workbook()
    bold = Font(bold=True)
    header_align = Alignment(vertical="center")

    def _write_sheet(ws, headers: list[str], rows: list[list]):
        ws.append(headers)
        for cell in ws[1]:
            cell.font = bold
            cell.alignment = header_align
        for r in rows:
            ws.append(r)
        # Best-effort column auto-width — Excel handles fine-tuning, this is
        # just so columns aren't comically narrow on first open.
        for col_idx, header in enumerate(headers, start=1):
            max_len = len(str(header))
            for r in rows:
                if col_idx - 1 < len(r):
                    v = r[col_idx - 1]
                    if v is not None:
                        max_len = max(max_len, len(str(v)))
            letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 60)

    ws_lines = wb.active
    ws_lines.title = "Lines"
    _write_sheet(ws_lines, LINES_HEADERS, build_lines_rows(buffer))

    ws_zones = wb.create_sheet("Zones")
    _write_sheet(ws_zones, ZONES_HEADERS, build_zones_rows(buffer, pairings))

    ws_summary = wb.create_sheet("Summary")
    _write_sheet(ws_summary, SUMMARY_HEADERS,
                 build_summary_rows(buffer, pairings, area_pairings))

    ws_hourly = wb.create_sheet("Hourly_Summary")
    _write_sheet(ws_hourly, HOURLY_HEADERS,
                 build_hourly_rows(buffer, pairings, area_pairings,
                                   metadata, buffer_started_at, export_time))

    ws_daily = wb.create_sheet("Daily_Summary")
    _write_sheet(ws_daily, DAILY_HEADERS,
                 build_daily_rows(buffer, pairings, area_pairings,
                                  metadata, buffer_started_at, export_time))

    ws_meta = wb.create_sheet("Metadata")
    _write_sheet(ws_meta, ["Key", "Value"],
                 build_metadata_kv(metadata, buffer, pairings))

    path = os.path.join(folder, "report.xlsx")
    wb.save(path)
    return path


def export_events(output_root: str, project_name: str,
                  buffer: Iterable[dict],
                  currently_tracked_ids: set[int] | None,
                  metadata: dict) -> tuple[str, int]:
    """Write a snapshot of ``buffer`` into a timestamped subfolder of
    ``output_root``. Returns ``(folder_path, event_count)``.

    Caller responsibilities: take the buffer snapshot before calling (we
    iterate once, but the contract is "we don't mutate"), and supply the
    set of currently-tracked IDs so orphan enters can be classified as
    still_inside vs lost_id.

    Output: a single ``report.xlsx`` workbook (multiple sheets — see
    :func:`write_xlsx`). If openpyxl isn't installed the function still
    creates the folder but returns event_count without a workbook.
    """
    snap = list(buffer)
    tracked = set(currently_tracked_ids or [])
    export_time = float(metadata.get("export_time", time.time()))

    timestamp_str = time.strftime("%Y%m%d_%H%M%S",
                                   time.localtime(export_time))
    folder = os.path.join(
        output_root,
        f"export_{_safe_name(project_name)}_{timestamp_str}")
    os.makedirs(folder, exist_ok=True)

    snap.sort(key=lambda e: float(e.get("timestamp", 0.0)))
    pairings = compute_pairings(snap, tracked, export_time)
    entrance_ids = set(metadata.get("entrance_line_ids") or [])
    exit_ids = set(metadata.get("exit_line_ids") or [])
    bidirectional_ids = set(metadata.get("bidirectional_line_ids") or [])
    area_pairings = compute_area_pairings(
        snap, entrance_ids, exit_ids, tracked, export_time,
        bidirectional_ids=bidirectional_ids)

    write_xlsx(folder, snap, pairings, metadata, area_pairings)

    return folder, len(snap)
